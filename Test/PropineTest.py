from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import os
import openpyxl
from selenium.webdriver.chrome.options import Options

chrome_options = Options()
chrome_options.add_argument("--headless") # Runs Chrome in headless mode.
chrome_options.add_argument('--no-sandbox') # Bypass OS security model
chrome_options.add_argument('--disable-gpu')  # applicable to windows os only
chrome_options.add_argument('start-maximized') #
chrome_options.add_argument('disable-infobars')
chrome_options.add_argument("--disable-extensions")

#driver = webdriver.Chrome(chrome_options=options, executable_path=r'C:\path\to\chromedriver.exe')

driver = webdriver.Chrome(("..\\drivers\\chromedriver.exe"), options=chrome_options)
driver.set_page_load_timeout("10")
baseURL = "https://vast-dawn-73245.herokuapp.com/"
#driver.maximize_window()
driver.implicitly_wait(3)
driver.get(baseURL)

my_path = "..\\data\\TestData.xlsx"
my_wb_obj = openpyxl.load_workbook(my_path)
my_sheet_obj = my_wb_obj.active
my_row = my_sheet_obj.max_row
my_col = my_sheet_obj.max_column

for i in range(2, my_row + 1):
   cell_obj = my_sheet_obj.cell(row = i, column = 1)
   print(cell_obj.value)
   time.sleep(3)
   driver.find_element_by_xpath("//input[@class='form-control']").send_keys(cell_obj.value)
   time.sleep(3)
   driver.find_element_by_xpath("//input[@class='btn btn-default']").click()
   result = driver.find_element_by_xpath("/html/body/div[2]/div/div[2]/div").text
   print(result)

   actual = my_sheet_obj.cell(row=i, column=3)
   actual.value = result

   expected = my_sheet_obj.cell(row=i, column=2)
   if expected.value == actual.value:
      status = my_sheet_obj.cell(row=i, column=4)
      status.value = "Pass"
   else:
      status = my_sheet_obj.cell(row=i, column=4)
      status.value = "Fail"

my_wb_obj.save("..\\output\\Result.xlsx")
time.sleep(5)
print('Execution Complete')
driver.quit()